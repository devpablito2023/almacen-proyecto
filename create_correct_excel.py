import openpyxl
from openpyxl.styles import Font

# Crear nuevo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

# Headers correctos que incluyen la columna Tipo
headers = [
    "Código", "Nombre", "Descripción", "Tipo", "Categoría", "Magnitud", 
    "Stock Mínimo", "Stock Máximo", "Stock Crítico", "Precio Unitario", 
    "Ubicación Física", "Estado", "Fecha Creación"
]

# Escribir headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

# Datos de ejemplo
productos = [
    ["TESTE2", "TESTE2", "TESTE2", "repuesto", "Materiales de Construcción", "GAL", 1, 100, 0, 0, "TESTE2", "Activo", "2025-08-26 17:01:11"],
    ["TESTE1", "TESTE1", "TESTE1", "insumo", "Materiales de Construcción", "UND", 1, 100, 0, 0, "TESTE1", "Inactivo", "2025-08-26 16:27:11"],
    ["STRING", "string", "string", "insumo", "string", "UND", 1, 100, 0, 0, "string", "Activo", "2025-08-26 15:54:34"]
]

# Escribir datos
for row_idx, producto in enumerate(productos, 2):
    for col_idx, value in enumerate(producto, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Guardar archivo
wb.save(r"C:\Users\Ztrack2\Desktop\almacen-proyecto\test_import_correcto.xlsx")
print("✅ Archivo Excel correcto creado: test_import_correcto.xlsx")

# Verificar headers
print("\n📋 Headers del archivo creado:")
for i, header in enumerate(headers, 1):
    print(f"{i}: {header}")
