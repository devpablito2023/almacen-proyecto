# backend/app/server/functions/productos.py
from fastapi import HTTPException, status
from server.config.database import (
    productos_collection,
    stock_collection, 
    get_next_id, 
    save_to_history, 
    log_activity
)
from server.models.productos import ProductoCreate, ProductoUpdate, ProductoSearch
from server.config.database import stats_collection

from datetime import datetime
from typing import Optional, List
import logging
import math

logger = logging.getLogger(__name__)

async def crear_producto(producto_data: ProductoCreate, created_by: int, created_by_name: str):
    """Crear nuevo producto"""
    try:
        # Verificar que código no exista
        existing_codigo = await productos_collection().find_one(
            {"codigo_producto": producto_data.codigo_producto}
        )
        if existing_codigo:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="El código de producto ya existe"
            )
        
        # Generar ID autoincremental
        nuevo_id = await get_next_id("productos")
        
        # Preparar datos del producto
        producto_dict = {
            "id_producto": nuevo_id,
            "codigo_producto": producto_data.codigo_producto,
            "nombre_producto": producto_data.nombre_producto,
            "tipo_producto": producto_data.tipo_producto,
            "categoria_producto": producto_data.categoria_producto,
            "proveedor_producto": producto_data.proveedor_producto,
            "costo_unitario": float(producto_data.costo_unitario) if producto_data.costo_unitario else None,
            "precio_referencial": float(producto_data.precio_referencial) if producto_data.precio_referencial else None,
            "ubicacion_fisica": producto_data.ubicacion_fisica,
            "stock_minimo": producto_data.stock_minimo,
            "stock_maximo": producto_data.stock_maximo,
            "stock_critico": producto_data.stock_critico,
            "estado_producto": 1,
            "descripcion_producto": producto_data.descripcion_producto,
            "url_foto_producto": None,
            "magnitud_producto": producto_data.magnitud_producto,
            "requiere_lote": producto_data.requiere_lote,
            "dias_vida_util": producto_data.dias_vida_util,
            "created_at": datetime.now(),
            "created_by": created_by,
            "created_by_name": created_by_name,
            "updated_at": None,
            "updated_by": None,
            "updated_by_name": None
        }
        
        # Insertar producto
        result = await productos_collection().insert_one(producto_dict)
        
        # Crear registro inicial en stock
        stock_inicial = {
            "id_stock": await get_next_id("stock"),
            "producto_id": nuevo_id,
            "producto_codigo": producto_data.codigo_producto,
            "producto_nombre": producto_data.nombre_producto,
            "cantidad_disponible": 0,
            "cantidad_reservada": 0,
            "cantidad_total": 0,
            "ubicacion_fisica": producto_data.ubicacion_fisica,
            "lote_serie": None,
            "fecha_vencimiento": None,
            "costo_promedio": float(producto_data.costo_unitario) if producto_data.costo_unitario else 0.0,
            "valor_inventario": 0.0,
            "fecha_ultimo_movimiento": None,
            "estado_stock": 1,
            "alerta_generada": False,
            "created_at": datetime.now(),
            "created_by": created_by,
            "created_by_name": created_by_name
        }
        
        await stock_collection().insert_one(stock_inicial)
        
        # Guardar en histórico
        await save_to_history(
            "productos",
            producto_dict,
            "CREATED",
            created_by,
            created_by_name
        )
        
        # Log de actividad
        await log_activity(
            action="PRODUCT_CREATED",
            module="productos",
            user_id=created_by,
            user_name=created_by_name,
            details={"product_id": nuevo_id, "codigo": producto_data.codigo_producto}
        )
        
        # Obtener producto creado
        producto_creado = await productos_collection().find_one(
            {"_id": result.inserted_id},
            {"_id": 0}
        )
        
        logger.info(f"Producto creado exitosamente: {producto_data.codigo_producto}")

        # 🔥 Actualizar estadísticas (se integra aquí)
        await update_stats_on_create(producto_creado)


        
        return producto_creado
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creando producto: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

async def obtener_producto_por_id(product_id: int):
    """Obtener producto por ID"""
    try:
        producto = await productos_collection().find_one(
            {"id_producto": product_id},
            {"_id": 0}
        )
        
        if not producto:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Producto no encontrado"
            )
        
        return producto
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo producto: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

async def obtener_productos(page: int = 1, limit: int = 20, estado: Optional[int] = None, tipo: Optional[str] = None):
    """Obtener lista de productos con paginación"""
    try:
        # Construir filtros
        filtros = {}
        if estado is not None:
            filtros["estado_producto"] = estado
        if tipo:
            filtros["tipo_producto"] = tipo
        
        # Calcular skip
        skip = (page - 1) * limit
        
        # Obtener total de registros
        total = await productos_collection().count_documents(filtros)
        
        # Obtener productos
        cursor = productos_collection().find(filtros, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
        productos = await cursor.to_list(length=limit)
        
        # Calcular paginación
        pages = math.ceil(total / limit) if total > 0 else 0

        #traer estadisticas
        stats = await obtener_estadistica()
        print(total)
        return {
            "data": {
                "productos" : productos,
                "stats": stats
            },

           
            "total": total,
            "page": page,
            "limit": limit,
            "pages": pages,
            "has_next": page < pages,
            "has_prev": page > 1
    
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo productos: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

async def actualizar_producto(product_id: int, producto_data: ProductoUpdate, updated_by: int, updated_by_name: str):
    """Actualizar producto"""
    try:
        # Verificar que producto existe
        producto_actual = await productos_collection().find_one({"id_producto": product_id})
        if not producto_actual:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Producto no encontrado"
            )
        
        # Preparar datos para actualizar
        update_data = {}
        for field, value in producto_data.dict(exclude_unset=True).items():
            if value is not None:
                if field in ["costo_unitario", "precio_referencial"] and value is not None:
                    update_data[field] = float(value)
                else:
                    update_data[field] = value
        
        if not update_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No hay datos para actualizar"
            )
        
        # Agregar metadatos de actualización
        update_data.update({
            "updated_at": datetime.now(),
            "updated_by": updated_by,
            "updated_by_name": updated_by_name
        })
        
        # Actualizar en BD
        result = await productos_collection().update_one(
            {"id_producto": product_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se realizaron cambios"
            )
        
        # Obtener producto actualizado
        producto_actualizado = await productos_collection().find_one(
            {"id_producto": product_id},
            {"_id": 0}
        )
        
        # Actualizar datos relacionados en stock si es necesario
        stock_update = {}
        if "nombre_producto" in update_data:
            stock_update["producto_nombre"] = update_data["nombre_producto"]
        if "ubicacion_fisica" in update_data:
            stock_update["ubicacion_fisica"] = update_data["ubicacion_fisica"]
        if "costo_unitario" in update_data:
            stock_update["costo_promedio"] = update_data["costo_unitario"]
        
        if stock_update:
            await stock_collection().update_one(
                {"producto_id": product_id},
                {"$set": stock_update}
            )
        
        # Guardar en histórico
        await save_to_history(
            "productos",
            producto_actualizado,
            "UPDATED",
            updated_by,
            updated_by_name
        )
        
        # Log de actividad
        await log_activity(
            action="PRODUCT_UPDATED",
            module="productos",
            user_id=updated_by,
            user_name=updated_by_name,
            details={"product_id": product_id, "fields": list(update_data.keys())}
        )
        
        logger.info(f"Producto actualizado: ID {product_id}")
        
        return producto_actualizado
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizando producto: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
       )

async def eliminar_producto(product_id: int, deleted_by: int, deleted_by_name: str):
   """Eliminar producto (soft delete)"""
   try:
       # Verificar que producto existe y no está ya eliminado
       producto = await productos_collection().find_one({
           "id_producto": product_id,
           "estado_producto": 1
       })
       
       if not producto:
           raise HTTPException(
               status_code=status.HTTP_404_NOT_FOUND,
               detail="Producto no encontrado o ya eliminado"
           )
       
       # Verificar que no tenga stock
       stock = await stock_collection().find_one({"producto_id": product_id})
       if stock and stock["cantidad_total"] > 0:
           raise HTTPException(
               status_code=status.HTTP_400_BAD_REQUEST,
               detail="No se puede eliminar un producto con stock existente"
           )
       
       # Soft delete
       result = await productos_collection().update_one(
           {"id_producto": product_id},
           {
               "$set": {
                   "estado_producto": 0,
                   "updated_at": datetime.now(),
                   "updated_by": deleted_by,
                   "updated_by_name": deleted_by_name
               }
           }
       )
       
       # Eliminar también el stock
       await stock_collection().update_one(
           {"producto_id": product_id},
           {
               "$set": {
                   "estado_stock": 0,
                   "updated_at": datetime.now()
               }
           }
       )
       
       # Guardar en histórico
       producto["estado_producto"] = 0
       await save_to_history(
           "productos",
           producto,
           "DELETED",
           deleted_by,
           deleted_by_name
       )
       
       # Log de actividad
       await log_activity(
           action="PRODUCT_DELETED",
           module="productos",
           user_id=deleted_by,
           user_name=deleted_by_name,
           details={"product_id": product_id, "codigo": producto["codigo_producto"]}
       )
       
       logger.info(f"Producto eliminado: ID {product_id}")
       
       return {"message": "Producto eliminado exitosamente"}
       
   except HTTPException:
       raise
   except Exception as e:
       logger.error(f"Error eliminando producto: {e}")
       raise HTTPException(
           status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
           detail="Error interno del servidor"
       )

async def restablecer_producto(product_id: int, restored_by: int, restored_by_name: str):
    """Restablecer producto eliminado (revertir soft delete)"""
    try:
        # Verificar que producto existe y está eliminado
        producto = await productos_collection().find_one({
            "id_producto": product_id,
            "estado_producto": 0
        })
        
        if not producto:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Producto no encontrado o ya está activo"
            )
        
        # Restablecer producto (soft restore)
        result = await productos_collection().update_one(
            {"id_producto": product_id},
            {
                "$set": {
                    "estado_producto": 1,
                    "updated_at": datetime.now(),
                    "updated_by": restored_by,
                    "updated_by_name": restored_by_name
                }
            }
        )
        
        # Restablecer también el stock
        await stock_collection().update_one(
            {"producto_id": product_id},
            {
                "$set": {
                    "estado_stock": 1,
                    "updated_at": datetime.now()
                }
            }
        )
        
        # Guardar en histórico
        producto["estado_producto"] = 1
        await save_to_history(
            "productos",
            producto,
            "RESTORED",
            restored_by,
            restored_by_name
        )
        
        # Log de actividad
        await log_activity(
            action="PRODUCT_RESTORED",
            module="productos",
            user_id=restored_by,
            user_name=restored_by_name,
            details={"product_id": product_id, "codigo": producto["codigo_producto"]}
        )
        
        logger.info(f"Producto restablecido: ID {product_id}")
        
        return {"message": "Producto restablecido exitosamente"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error restableciendo producto: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

async def buscar_productos(search_params: ProductoSearch, limit: int = 20):
   """Buscar productos con filtros"""
   try:
       # Construir filtros de búsqueda
       filtros = {}
       
       if search_params.codigo:
           filtros["codigo_producto"] = {"$regex": search_params.codigo, "$options": "i"}
       
       if search_params.nombre:
           filtros["nombre_producto"] = {"$regex": search_params.nombre, "$options": "i"}
       
       if search_params.tipo:
           filtros["tipo_producto"] = search_params.tipo
       
       if search_params.categoria:
           filtros["categoria_producto"] = {"$regex": search_params.categoria, "$options": "i"}
       
       if search_params.proveedor:
           filtros["proveedor_producto"] = {"$regex": search_params.proveedor, "$options": "i"}
       
       if search_params.estado is not None:
           filtros["estado_producto"] = search_params.estado
       
       # Si busca productos con stock bajo, hacer join con stock
       if search_params.stock_bajo:
           pipeline = [
               {"$match": filtros},
               {
                   "$lookup": {
                       "from": "stock",
                       "localField": "id_producto",
                       "foreignField": "producto_id",
                       "as": "stock_info"
                   }
               },
               {"$unwind": "$stock_info"},
               {
                   "$match": {
                       "$expr": {
                           "$lte": ["$stock_info.cantidad_total", "$stock_minimo"]
                       }
                   }
               },
               {"$project": {"_id": 0, "stock_info": 0}},
               {"$limit": limit}
           ]
           
           cursor = productos_collection().aggregate(pipeline)
           productos = await cursor.to_list(length=limit)
       else:
           # Búsqueda normal
           cursor = productos_collection().find(filtros, {"_id": 0}).sort("nombre_producto", 1).limit(limit)
           productos = await cursor.to_list(length=limit)
       
       return productos
       
   except Exception as e:
       logger.error(f"Error buscando productos: {e}")
       raise HTTPException(
           status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
           detail="Error interno del servidor"
       )

async def verificar_codigo_producto_unico(codigo: str, exclude_id: Optional[int] = None):
   """Verificar si un código de producto es único"""
   try:
       filtros = {"codigo_producto": codigo}
       if exclude_id:
           filtros["id_producto"] = {"$ne": exclude_id}
       
       existing = await productos_collection().find_one(filtros)
       
       return existing is None
       
   except Exception as e:
       logger.error(f"Error verificando código único: {e}")
       return False

async def obtener_productos_autocomplete(query: str, limit: int = 10):
   """Obtener productos para autocomplete"""
   try:
       filtros = {
           "estado_producto": 1,
           "$or": [
               {"codigo_producto": {"$regex": query, "$options": "i"}},
               {"nombre_producto": {"$regex": query, "$options": "i"}}
           ]
       }
       
       cursor = productos_collection().find(
           filtros,
           {"id_producto": 1, "codigo_producto": 1, "nombre_producto": 1, "magnitud_producto": 1, "_id": 0}
       ).sort("nombre_producto", 1).limit(limit)
       
       productos = await cursor.to_list(length=limit)
       
       return productos
       
   except Exception as e:
       logger.error(f"Error en autocomplete de productos: {e}")
       raise HTTPException(
           status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
           detail="Error interno del servidor"
       )

async def update_stats_on_create(producto):
    await stats_collection().update_one(
        {"_id": "productos_stats"},
        {
            "$inc": {
                "total_productos": 1,
                "productos_activos": 1,
                f"por_tipo.{producto['tipo_producto']}": 1,
                f"por_categoria.{producto['categoria_producto']}": 1,
            },
            "$set": {"last_updated": datetime.now()}
        },
        upsert=True
    )

async def update_stats_on_delete(producto):
    await stats_collection().update_one(
        {"_id": "productos_stats"},
        {
            "$inc": {
                "total_productos": -1,
                "productos_activos": -1 if producto["estado_producto"] == 1 else 0,
                f"por_tipo.{producto['tipo_producto']}": -1,
                f"por_categoria.{producto['categoria_producto']}": -1,
            },
            "$set": {"last_updated": datetime.now()}
        }
    )


async def update_stats_on_stock_change(stock, old_total, new_total):
    updates = {}
    if old_total == 0 and new_total > 0:
        updates["productos_sin_stock"] = -1
    elif old_total > 0 and new_total == 0:
        updates["productos_sin_stock"] = 1
    
    if updates:
        await stats_collection().update_one(
            {"_id": "productos_stats"},
            {
                "$inc": updates,
                "$set": {"last_updated": datetime.now()}
            }
        )

async def obtener_estadistica():
    # Calcular estadísticas directamente
    stats = await obtener_estadisticas_productos()
    return stats


async def actualizar_estadistica():
    stats = await obtener_estadisticas_productos()  # tu función de antes
    stats["_id"] = "productos_stats"
    await stats_collection().replace_one({"_id": "productos_stats"}, stats, upsert=True)

async def obtener_estadisticas_productos():
    """Calcular estadísticas globales de productos para dashboard"""
    try:
        # Total productos
        total_productos = await productos_collection().count_documents({})
        
        # Productos activos
        productos_activos = await productos_collection().count_documents({"estado_producto": 1})
        
        # Productos sin stock
        productos_sin_stock = await stock_collection().count_documents({"cantidad_total": 0})
        
        # Productos con stock bajo (<= stock_minimo)
        pipeline_bajo = [
            {"$lookup": {
                "from": "stock",
                "localField": "id_producto",
                "foreignField": "producto_id",
                "as": "stock_info"
            }},
            {"$unwind": "$stock_info"},
            {"$match": {"$expr": {"$lte": ["$stock_info.cantidad_total", "$stock_minimo"]}}}
        ]
        productos_stock_bajo = len(await productos_collection().aggregate(pipeline_bajo).to_list(length=None))
        
        # Productos con stock crítico (<= stock_critico)
        pipeline_critico = [
            {"$lookup": {
                "from": "stock",
                "localField": "id_producto",
                "foreignField": "producto_id",
                "as": "stock_info"
            }},
            {"$unwind": "$stock_info"},
            {"$match": {"$expr": {"$lte": ["$stock_info.cantidad_total", "$stock_critico"]}}}
        ]
        productos_stock_critico = len(await productos_collection().aggregate(pipeline_critico).to_list(length=None))
        
        # Productos próximos a vencer (30 días)
        from datetime import datetime, timedelta
        fecha_limite = datetime.now() + timedelta(days=30)
        productos_proximos_vencer = await stock_collection().count_documents({
            "fecha_vencimiento": {"$lte": fecha_limite, "$ne": None}
        })
        
        # Valor total de inventario basado en productos (suma costos unitarios, null como 0)
        agg_valor = await productos_collection().aggregate([
            {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$costo_unitario", 0]}}}}
        ]).to_list(length=1)
        valor_total_inventario = agg_valor[0]["total"] if agg_valor else 0.0
        
        # Agrupación por tipo
        agg_tipo = await productos_collection().aggregate([
            {"$group": {"_id": "$tipo_producto", "count": {"$sum": 1}}}
        ]).to_list(length=None)
        por_tipo = {item["_id"]: item["count"] for item in agg_tipo}
        
        # Agrupación por categoría
        agg_categoria = await productos_collection().aggregate([
            {"$group": {"_id": "$categoria_producto", "count": {"$sum": 1}}}
        ]).to_list(length=None)
        por_categoria = {item["_id"]: item["count"] for item in agg_categoria}
        
        return {
            "total_productos": total_productos,
            "productos_activos": productos_activos,
            "productos_stock_bajo": productos_stock_bajo,
            "productos_stock_critico": productos_stock_critico,
            "productos_proximos_vencer": productos_proximos_vencer,
            "productos_sin_stock": productos_sin_stock,
            "valor_total_inventario": valor_total_inventario,
        }
    except Exception as e:
        logger.error(f"Error calculando estadísticas de productos: {e}")
        raise HTTPException(status_code=500, detail="Error interno al calcular estadísticas")

async def importar_productos_masivo(file_content: bytes, created_by: int, created_by_name: str):
    """Importar productos masivamente desde archivo Excel"""
    import openpyxl
    from io import BytesIO
    from server.models.productos import ProductoImportRow, ImportResult, ImportError
    
    try:
        # Cargar workbook desde bytes
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        worksheet = workbook.active
        
        productos_creados = []
        errores_detalle = []
        total_procesados = 0
        exitosos = 0
        errores = 0
        
        # Obtener headers de la primera fila
        headers = []
        for col in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=col).value
            if header_value:
                # Normalizar header: minúsculas, sin espacios ni acentos
                normalized_header = str(header_value).lower().strip()
                normalized_header = normalized_header.replace(' ', '_')
                normalized_header = normalized_header.replace('ó', 'o').replace('í', 'i').replace('á', 'a').replace('é', 'e').replace('ú', 'u')
                headers.append(normalized_header)
            else:
                headers.append(f"columna_{col}")
        
        # Mapeo de columnas esperadas
        column_mapping = {
            'codigo_producto': ['codigo', 'codigo_producto', 'code', 'sku'],
            'nombre_producto': ['nombre', 'nombre_producto', 'name', 'product_name'],
            'tipo_producto': ['tipo', 'tipo_producto', 'type', 'product_type'],
            'categoria_producto': ['categoria', 'categoria_producto', 'category'],
            'proveedor_producto': ['proveedor', 'proveedor_producto', 'supplier'],
            'costo_unitario': ['costo', 'costo_unitario', 'cost', 'unit_cost'],
            'precio_referencial': ['precio', 'precio_referencial', 'price', 'reference_price', 'precio_unitario'],
            'ubicacion_fisica': ['ubicacion', 'ubicacion_fisica', 'location', 'ubicacion_fisica'],
            'stock_minimo': ['stock_min', 'stock_minimo', 'min_stock', 'stock_minimo'],
            'stock_maximo': ['stock_max', 'stock_maximo', 'max_stock', 'stock_maximo'],
            'stock_critico': ['stock_critico', 'critical_stock', 'stock_critico'],
            'descripcion_producto': ['descripcion', 'descripcion_producto', 'description'],
            'magnitud_producto': ['magnitud', 'magnitud_producto', 'unit', 'unidad'],
            'requiere_lote': ['requiere_lote', 'batch_required', 'lote'],
            'dias_vida_util': ['dias_vida_util', 'shelf_life', 'vida_util']
        }
        
        # Encontrar índices de columnas
        column_indices = {}
        for field, possible_names in column_mapping.items():
            for i, header in enumerate(headers):
                if header in possible_names:
                    column_indices[field] = i
                    break
        
        # Verificar columnas requeridas
        required_columns = ['codigo_producto', 'nombre_producto', 'tipo_producto']
        missing_columns = [col for col in required_columns if col not in column_indices]
        
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Columnas requeridas faltantes: {', '.join(missing_columns)}"
            )
        
        # Procesar filas (empezar desde la fila 2)
        for row_num in range(2, worksheet.max_row + 1):
            total_procesados += 1
            row_errors = []
            
            try:
                # Extraer datos de la fila
                row_data = {}
                for field, col_index in column_indices.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    
                    if cell_value is not None:
                        # Limpiar y convertir valores
                        if field in ['costo_unitario', 'precio_referencial']:
                            try:
                                row_data[field] = float(str(cell_value).replace(',', '.'))
                            except (ValueError, TypeError):
                                row_data[field] = None
                        elif field in ['stock_minimo', 'stock_maximo', 'stock_critico', 'dias_vida_util']:
                            try:
                                row_data[field] = int(float(str(cell_value)))
                            except (ValueError, TypeError):
                                row_data[field] = 1 if field in ['stock_minimo', 'stock_maximo'] else 0
                        elif field == 'requiere_lote':
                            str_value = str(cell_value).upper().strip()
                            row_data[field] = str_value in ['SI', 'SÍ', 'YES', 'TRUE', '1', 'VERDADERO']
                        else:
                            row_data[field] = str(cell_value).strip()
                    else:
                        # Valores por defecto
                        if field in ['stock_minimo', 'stock_maximo']:
                            row_data[field] = 1 if field == 'stock_minimo' else 100
                        elif field == 'stock_critico':
                            row_data[field] = 0
                        elif field == 'magnitud_producto':
                            row_data[field] = 'UND'
                        elif field == 'requiere_lote':
                            row_data[field] = False
                        else:
                            row_data[field] = None
                
                # Validar datos requeridos
                if not row_data.get('codigo_producto'):
                    row_errors.append("Código de producto es requerido")
                if not row_data.get('nombre_producto'):
                    row_errors.append("Nombre de producto es requerido")
                if not row_data.get('tipo_producto'):
                    row_errors.append("Tipo de producto es requerido")
                
                # Validar tipo de producto
                if row_data.get('tipo_producto'):
                    allowed_types = ['insumo', 'repuesto', 'herramienta', 'otro']
                    if row_data['tipo_producto'].lower() not in allowed_types:
                        row_errors.append(f"Tipo debe ser uno de: {', '.join(allowed_types)}")
                    else:
                        row_data['tipo_producto'] = row_data['tipo_producto'].lower()
                
                # Verificar código único
                if row_data.get('codigo_producto'):
                    existing_codigo = await productos_collection().find_one(
                        {"codigo_producto": row_data['codigo_producto'].upper()}
                    )
                    if existing_codigo:
                        row_errors.append(f"El código '{row_data['codigo_producto']}' ya existe")
                    else:
                        row_data['codigo_producto'] = row_data['codigo_producto'].upper()
                
                # Si hay errores, agregar a la lista de errores
                if row_errors:
                    errores += 1
                    errores_detalle.append(ImportError(
                        fila=row_num,
                        errores=row_errors,
                        datos=row_data
                    ))
                    continue
                
                # Crear ProductoCreate object
                producto_create = ProductoCreate(
                    codigo_producto=row_data['codigo_producto'],
                    nombre_producto=row_data['nombre_producto'],
                    tipo_producto=row_data['tipo_producto'],
                    categoria_producto=row_data.get('categoria_producto'),
                    proveedor_producto=row_data.get('proveedor_producto'),
                    costo_unitario=row_data.get('costo_unitario'),
                    precio_referencial=row_data.get('precio_referencial'),
                    ubicacion_fisica=row_data.get('ubicacion_fisica'),
                    stock_minimo=row_data.get('stock_minimo', 1),
                    stock_maximo=row_data.get('stock_maximo', 100),
                    stock_critico=row_data.get('stock_critico', 0),
                    descripcion_producto=row_data.get('descripcion_producto'),
                    magnitud_producto=row_data.get('magnitud_producto', 'UND'),
                    requiere_lote=row_data.get('requiere_lote', False),
                    dias_vida_util=row_data.get('dias_vida_util')
                )
                
                # Crear producto
                producto_creado = await crear_producto(
                    producto_create,
                    created_by=created_by,
                    created_by_name=created_by_name
                )
                
                productos_creados.append(producto_creado)
                exitosos += 1
                
            except Exception as e:
                errores += 1
                error_msg = str(e)
                if "El código de producto ya existe" in error_msg:
                    error_msg = f"Código duplicado: {row_data.get('codigo_producto', 'N/A')}"
                
                errores_detalle.append(ImportError(
                    fila=row_num,
                    errores=[error_msg],
                    datos=row_data if 'row_data' in locals() else {}
                ))
                logger.error(f"Error procesando fila {row_num}: {e}")
        
        # Crear resultado
        resultado = ImportResult(
            total_procesados=total_procesados,
            exitosos=exitosos,
            errores=errores,
            productos_creados=productos_creados,
            errores_detalle=errores_detalle
        )
        
        return resultado
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en importación masiva: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error procesando archivo Excel: {str(e)}"
        )


