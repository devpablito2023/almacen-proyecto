# backend/app/server/routes/stock.py
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from server.functions.stock import (
    obtener_stock,
    obtener_stock_por_producto,
    ajustar_stock,
    obtener_alertas_stock,
    calcular_valoracion_inventario,
    actualizar_stock_stats,
    obtener_estadistica_stock,
    obtener_movimientos_stock
)
from server.models.stock import StockAdjust
from server.models.responses import success_response, error_response, paginated_response
from server.routes.auth import get_current_user
from server.config.security import check_permission
from typing import Optional
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter()

def check_stock_permission(user_type: int, action: str): 
    """Verificar permisos de usuario para módulo stock"""
    if not check_permission(user_type, "stock", action):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tiene permisos para esta operación"
        )

@router.get("/stats_ok")
async def actualizar_stock_stats_ok():
    stats = await actualizar_stock_stats()
    if not stats:
        return {"message": "No procesado"}
    return stats

@router.get("/stats")
async def obtener_estadistica_stock_ok():
    #stats = await stats_collection().find_one({"_id": "productos_stats"}, {"_id": 0})
    stats = await obtener_estadistica_stock()
    if not stats:
        return {"message": "No hay estadísticas registradas"}
    return stats


@router.get("/", summary="Consultar stock general")
async def get_stock(
    page: int = Query(1, ge=1, description="Número de página"),
    limit: int = Query(20, ge=1, le=100, description="Elementos por página"),
    stock_bajo: bool = Query(False, description="Solo productos con stock bajo"),
    stock_critico: bool = Query(False, description="Solo productos con stock crítico"),
    current_user = Depends(get_current_user)
):
    """
    Obtener estado general del stock con paginación
    
    - **page**: Número de página (default: 1)
    - **limit**: Elementos por página (default: 20, max: 100)
    - **stock_bajo**: Filtrar solo productos con stock bajo
    - **stock_critico**: Filtrar solo productos con stock crítico
    
    Requiere permisos de lectura de stock
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        result = await obtener_stock(page, limit, stock_bajo, stock_critico)
        
        return paginated_response(
            data=result["data"],
            total=result["total"],
            page=result["page"],
            limit=result["limit"],
            message="Stock obtenido exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="STOCK_QUERY_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error consultando stock: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/producto/{product_id}", summary="Consultar stock de producto específico")
async def get_stock_producto(
    product_id: int,
    current_user = Depends(get_current_user)
):
    """
    Obtener stock de un producto específico
    
    - **product_id**: ID del producto a consultar
    
    Requiere permisos de lectura de stock
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        result = await obtener_stock_por_producto(product_id)
        
        return success_response(
            data=result,
            message="Stock del producto obtenido exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_STOCK_NOT_FOUND",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error consultando stock del producto {product_id}: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.post("/ajustar", summary="Realizar ajuste de stock")
async def adjust_stock(
    adjustment_data: StockAdjust,
    current_user = Depends(get_current_user)
):
    """
    Realizar ajuste de stock (aumento o disminución)
    
    - **producto_id**: ID del producto
    - **cantidad_ajuste**: Cantidad a ajustar (positiva o negativa)
    - **motivo**: Motivo del ajuste
    - **ubicacion**: Ubicación física (opcional)
    - **lote_serie**: Lote o serie (opcional)
    
    Requiere permisos de actualización de stock
    """
    try:
        # DEBUG: Log de entrada
        logger.info(f"🔧 DEBUG - Ajuste de stock iniciado")
        logger.info(f"🔧 DEBUG - Datos recibidos: {adjustment_data}")
        logger.info(f"🔧 DEBUG - Usuario actual: {current_user['user']['nombre_usuario']}")
        
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        logger.info(f"🔧 DEBUG - Tipo de usuario: {user_type}")
        check_stock_permission(user_type, "update")
        logger.info(f"🔧 DEBUG - Permisos verificados exitosamente")
        
        result = await ajustar_stock(
            adjustment_data,
            adjusted_by=current_user["user"]["id_usuario"],
            adjusted_by_name=current_user["user"]["nombre_usuario"]
        )
        
        logger.info(f"🔧 DEBUG - Resultado del ajuste: {result}")
        
        return success_response(
            data=result,
            message="Ajuste de stock realizado exitosamente"
        )
        
    except HTTPException as e:
        logger.error(f"🔧 DEBUG - HTTPException en ajuste: {e.detail}")
        return error_response(
            error="STOCK_ADJUSTMENT_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"🔧 DEBUG - Error inesperado en ajuste: {e}")
        logger.error(f"Error ajustando stock: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/alertas", summary="Obtener alertas de stock")
async def get_alertas_stock(
    current_user = Depends(get_current_user)
):
    """
    Obtener alertas de stock bajo, crítico y vencimientos
    
    Requiere permisos de lectura de stock
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        alertas = await obtener_alertas_stock()
        
        # Agrupar alertas por tipo
        alertas_agrupadas = {
            "critico": [a for a in alertas if a.tipo_alerta == "critico"],
            "bajo": [a for a in alertas if a.tipo_alerta == "bajo"],
            "vencimiento": [a for a in alertas if a.tipo_alerta == "vencimiento"]
        }
        
        # Convertir a dict para serialización
        result = {
            "total_alertas": len(alertas),
            "alertas_criticas": len(alertas_agrupadas["critico"]),
            "alertas_stock_bajo": len(alertas_agrupadas["bajo"]),
            "alertas_vencimiento": len(alertas_agrupadas["vencimiento"]),
            "alertas": [alert.dict() for alert in alertas],
            "alertas_agrupadas": {
                k: [alert.dict() for alert in v] 
                for k, v in alertas_agrupadas.items()
            }
        }
        
        return success_response(
            data=result,
            message=f"Se encontraron {len(alertas)} alertas"
        )
        
    except HTTPException as e:
        return error_response(
            error="ALERTS_QUERY_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error obteniendo alertas: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/valoracion", summary="Obtener valorización del inventario")
async def get_valoracion_inventario(
    current_user = Depends(get_current_user)
):
    """
    Obtener valorización completa del inventario
    
    Incluye:
    - Valor total del inventario
    - Cantidad de productos
    - Estadísticas de stock
    - Productos vencidos y por vencer
    
    Requiere permisos de lectura de stock
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        valoracion = await calcular_valoracion_inventario()
        
        return success_response(
            data=valoracion.dict(),
            message="Valorización calculada exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="VALUATION_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error calculando valorización: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/movimientos", summary="Obtener movimientos de stock")
async def get_movimientos_stock(
    producto_id: Optional[int] = Query(None, description="ID del producto específico"),
    limit: int = Query(50, ge=1, le=200, description="Límite de registros"),
    current_user = Depends(get_current_user)
):
    """
    Obtener histórico de movimientos de stock
    
    - **producto_id**: Filtrar por producto específico (opcional)
    - **limit**: Límite de registros (default: 50, max: 200)
    
    Requiere permisos de lectura de stock
    
    Nota: En Fase 1 retorna lista vacía. Se implementará en Fase 2 con kardex.
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        # TODO: Implementar cuando se tenga kardex en Fase 2
        movimientos = await obtener_movimientos_stock(producto_id, limit)
        
        return success_response(
            data=movimientos,
            message="Movimientos obtenidos exitosamente (funcionalidad completa en Fase 2)"
        )
        
    except HTTPException as e:
        return error_response(
            error="MOVEMENTS_QUERY_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error obteniendo movimientos: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/resumen", summary="Obtener resumen del stock")
async def get_resumen_stock(
    current_user = Depends(get_current_user)
):
    """
    Obtener resumen ejecutivo del stock
    
    Combina:
    - Valorización del inventario
    - Alertas activas
    - Estadísticas generales
    
    Requiere permisos de lectura de stock
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_stock_permission(user_type, "read")
        
        # Obtener valorización
        valoracion = await calcular_valoracion_inventario()
        
        # Obtener alertas
        alertas = await obtener_alertas_stock()
        
        # Construir resumen
        resumen = {
            "valoracion": valoracion.dict(),
            "alertas": {
                "total": len(alertas),
                "criticas": len([a for a in alertas if a.urgencia == "critica"]),
                "altas": len([a for a in alertas if a.urgencia == "alta"]),
                "medias": len([a for a in alertas if a.urgencia == "media"])
            },
            "indicadores": {
                "productos_activos": valoracion.total_productos,
                "valor_promedio_producto": (
                    valoracion.valor_total_inventario / valoracion.total_productos
                    if valoracion.total_productos > 0 else 0
                ),
                "porcentaje_stock_bajo": (
                    (valoracion.productos_stock_bajo / valoracion.total_productos * 100)
                    if valoracion.total_productos > 0 else 0
                ),
                "porcentaje_stock_critico": (
                    (valoracion.productos_stock_critico / valoracion.total_productos * 100)
                    if valoracion.total_productos > 0 else 0
                )
            }
        }
        
        return success_response(
            data=resumen,
            message="Resumen de stock obtenido exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="SUMMARY_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error obteniendo resumen: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/export", summary="Exportar stock a Excel")
async def exportar_stock(
    current_user=Depends(get_current_user),
    stock_bajo: Optional[bool] = Query(False, description="Filtro por stock bajo"),
    stock_critico: Optional[bool] = Query(False, description="Filtro por stock crítico"),
):
    """
    Exportar stock a archivo Excel con filtros aplicados
    """
    try:
        # Debug: ver qué parámetros llegan
        logger.info(f"DEBUG STOCK EXPORT - stock_bajo: {stock_bajo}, stock_critico: {stock_critico}")
        logger.info(f"DEBUG STOCK EXPORT - user_type: {current_user['user']['tipo_usuario']}")
        
        # Verificar permisos
        check_stock_permission(current_user["user"]["tipo_usuario"], "read")
        
        # Obtener stock con filtros
        stock_data = await obtener_stock(
            page=1,
            limit=10000,  # Obtener todo el stock
            stock_bajo=stock_bajo,
            stock_critico=stock_critico
        )
        
        stock_items = stock_data["data"]["stock"]
        
        # Crear workbook y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock"
        
        # Definir headers
        headers = [
            "Código Producto", "Nombre Producto", "Cantidad Disponible", "Cantidad Reservada", 
            "Cantidad Total", "Stock Mínimo", "Stock Crítico", "Magnitud",
            "Ubicación Física", "Lote/Serie", "Costo Promedio", "Valor Inventario",
            "Fecha Vencimiento", "Último Movimiento", "Alerta"
        ]
        
        # Aplicar estilo a headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        for row, item in enumerate(stock_items, 2):
            ws.cell(row=row, column=1, value=item.get("producto_codigo", ""))
            ws.cell(row=row, column=2, value=item.get("producto_nombre", ""))
            ws.cell(row=row, column=3, value=item.get("cantidad_disponible", 0))
            ws.cell(row=row, column=4, value=item.get("cantidad_reservada", 0))
            ws.cell(row=row, column=5, value=item.get("cantidad_total", 0))
            ws.cell(row=row, column=6, value=item.get("stock_minimo", 0))
            ws.cell(row=row, column=7, value=item.get("stock_critico", 0))
            ws.cell(row=row, column=8, value=item.get("magnitud", ""))
            ws.cell(row=row, column=9, value=item.get("ubicacion_fisica", ""))
            ws.cell(row=row, column=10, value=item.get("lote_serie", ""))
            ws.cell(row=row, column=11, value=item.get("costo_promedio", 0))
            ws.cell(row=row, column=12, value=item.get("valor_inventario", 0))
            ws.cell(row=row, column=13, value=item.get("fecha_vencimiento", ""))
            ws.cell(row=row, column=14, value=item.get("fecha_ultimo_movimiento", ""))
            ws.cell(row=row, column=15, value="Sí" if item.get("alerta_generada", False) else "No")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_export_{timestamp}.xlsx"
        
        logger.info(f"Exportación de stock completada por usuario {current_user['user']['codigo_usuario']}")
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exportando stock: {e}")
        return error_response(
            error="EXPORT_ERROR",
            message="Error al exportar stock",
            code=500
        )