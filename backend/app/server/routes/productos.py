# backend/app/server/routes/productos.py
from fastapi import APIRouter, Depends, HTTPException, status, Query, File, UploadFile
from fastapi.responses import StreamingResponse
from server.functions.productos import (
    crear_producto,
    obtener_producto_por_id,
    obtener_productos,
    actualizar_producto,
    eliminar_producto,
    restablecer_producto,
    buscar_productos,
    verificar_codigo_producto_unico,
    obtener_estadistica,
    actualizar_estadistica,
    obtener_productos_autocomplete,
    importar_productos_masivo
)
from server.models.productos import ProductoCreate, ProductoUpdate, ProductoSearch, ImportResult
from server.models.responses import success_response, error_response, paginated_response
from server.routes.auth import get_current_user
from server.config.security import check_permission
from typing import Optional
import logging
import os
import uuid
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter()

@router.get("/stats_ok")
async def actualizar_estadistica_productos():
    stats = await actualizar_estadistica()
    if not stats:
        return {"message": "No procesado"}
    return stats

@router.get("/stats")
async def obtener_estadisticas_productos():
    #stats = await stats_collection().find_one({"_id": "productos_stats"}, {"_id": 0})
    stats = await obtener_estadistica()
    if not stats:
        return {"message": "No hay estadísticas registradas"}
    return stats

def check_product_permission(user_type: int, action: str):
    """Verificar permisos de usuario para módulo productos"""
    if not check_permission(user_type, "productos", action):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tiene permisos para esta operación"
        )

@router.post("/", summary="Crear nuevo producto")
async def create_producto(
    producto_data: ProductoCreate,
    current_user = Depends(get_current_user)
):
    """
    Crear un nuevo producto en el sistema
    
    - **codigo_producto**: Código único del producto
    - **nombre_producto**: Nombre del producto
    - **tipo_producto**: Tipo (insumo, repuesto, herramienta, otro)
    - **categoria_producto**: Categoría del producto
    - **proveedor_producto**: Proveedor principal
    - **costo_unitario**: Costo unitario
    - **stock_minimo**: Stock mínimo
    - **stock_maximo**: Stock máximo
    - **stock_critico**: Stock crítico
    
    Requiere permisos de creación de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "create")
        
        result = await crear_producto(
            producto_data,
            created_by=current_user["user"]["id_usuario"],
            created_by_name=current_user["user"]["nombre_usuario"]
        )
        
        return success_response(
            data=result,
            message="Producto creado exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_CREATION_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error creando producto: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/", summary="Listar productos")
async def list_productos(
    page: int = Query(1, ge=1, description="Número de página"),
    limit: int = Query(20, ge=1, le=100, description="Elementos por página"),
    estado: Optional[int] = Query(None, ge=0, le=1, description="Estado del producto"),
    tipo: Optional[str] = Query(None, description="Tipo de producto"),
    current_user = Depends(get_current_user)
):
    """
    Obtener lista paginada de productos
    
    - **page**: Número de página (default: 1)
    - **limit**: Elementos por página (default: 20, max: 100)
    - **estado**: Filtrar por estado (0=inactivo, 1=activo)
    - **tipo**: Filtrar por tipo de producto
    
    Requiere permisos de lectura de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "read")
        
        result = await obtener_productos(page, limit, estado, tipo)
        
        return paginated_response(
            data=result["data"],
            total=result["total"],
            page=result["page"],
            limit=result["limit"],
            message="Productos obtenidos exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_LIST_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error listando productos: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/autocomplete", summary="Autocompletar productos")
async def autocomplete_productos(
    q: str = Query(..., min_length=1, description="Término de búsqueda"),
    limit: int = Query(10, ge=1, le=50, description="Límite de resultados"),
    current_user = Depends(get_current_user)
):
    """
    Obtener productos para autocompletar
    
    - **q**: Término de búsqueda
    - **limit**: Límite de resultados (default: 10, max: 50)
    
    Requiere permisos de lectura de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "read")
        
        result = await obtener_productos_autocomplete(q, limit)
        
        return success_response(
            data=result,
            message=f"Se encontraron {len(result)} productos"
        )
        
    except HTTPException as e:
        return error_response(
            error="AUTOCOMPLETE_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error en autocomplete: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/search", summary="Buscar productos")
async def search_productos(
    codigo: Optional[str] = Query(None, description="Código del producto"),
    nombre: Optional[str] = Query(None, description="Nombre del producto"),
    tipo: Optional[str] = Query(None, description="Tipo de producto"),
    categoria: Optional[str] = Query(None, description="Categoría"),
    proveedor: Optional[str] = Query(None, description="Proveedor"),
    estado: Optional[int] = Query(None, ge=0, le=1, description="Estado"),
    stock_bajo: Optional[bool] = Query(False, description="Solo productos con stock bajo"),
    limit: int = Query(20, ge=1, le=100, description="Límite de resultados"),
    current_user = Depends(get_current_user)
):
    """
    Buscar productos con filtros avanzados
    
    - **codigo**: Filtrar por código
    - **nombre**: Filtrar por nombre
    - **tipo**: Filtrar por tipo
    - **categoria**: Filtrar por categoría
    - **proveedor**: Filtrar por proveedor
    - **estado**: Filtrar por estado
    - **stock_bajo**: Solo productos con stock bajo
    - **limit**: Límite de resultados
    
    Requiere permisos de lectura de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "read")
        
        search_params = ProductoSearch(
            codigo=codigo,
            nombre=nombre,
            tipo=tipo,
            categoria=categoria,
            proveedor=proveedor,
            estado=estado,
            stock_bajo=stock_bajo
        )
        
        result = await buscar_productos(search_params, limit)
        
        return success_response(
            data=result,
            message=f"Se encontraron {len(result)} productos"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_SEARCH_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error buscando productos: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
       )

@router.get("/{product_id}", summary="Obtener producto por ID")
async def get_producto(
    product_id: int,
    current_user = Depends(get_current_user)
):
    """
    Obtener información de un producto específico
    
    - **product_id**: ID del producto a consultar
    
    Requiere permisos de lectura de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "read")
        
        result = await obtener_producto_por_id(product_id)
        
        return success_response(
            data=result,
            message="Producto obtenido exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_NOT_FOUND",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error obteniendo producto {product_id}: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.put("/{product_id}", summary="Actualizar producto")
async def update_producto(
    product_id: int,
    producto_data: ProductoUpdate,
    current_user = Depends(get_current_user)
):
    """
    Actualizar información de un producto
    
    - **product_id**: ID del producto a actualizar
    - **producto_data**: Datos a actualizar
    
    Requiere permisos de actualización de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "update")
        
        result = await actualizar_producto(
            product_id,
            producto_data,
            updated_by=current_user["user"]["id_usuario"],
            updated_by_name=current_user["user"]["nombre_usuario"]
        )
        
        return success_response(
            data=result,
            message="Producto actualizado exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_UPDATE_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error actualizando producto {product_id}: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.delete("/{product_id}", summary="Eliminar producto")
async def delete_producto(
    product_id: int,
    current_user = Depends(get_current_user)
):
    """
    Eliminar un producto (soft delete)
    
    - **product_id**: ID del producto a eliminar
    
    Requiere permisos de eliminación de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "delete")
        
        result = await eliminar_producto(
            product_id,
            deleted_by=current_user["user"]["id_usuario"],
            deleted_by_name=current_user["user"]["nombre_usuario"]
        )
        
        return success_response(
            data=result,
            message="Producto eliminado exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_DELETE_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error eliminando producto {product_id}: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.patch("/{product_id}/restore", summary="Restablecer producto")
async def restore_producto(
    product_id: int,
    current_user = Depends(get_current_user)
):
    """
    Restablecer un producto eliminado (revertir soft delete)
    
    - **product_id**: ID del producto a restablecer
    
    Requiere permisos de eliminación de productos
    """
    try:
        # Verificar permisos (mismo permiso que eliminar)
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "delete")
        
        result = await restablecer_producto(
            product_id,
            restored_by=current_user["user"]["id_usuario"],
            restored_by_name=current_user["user"]["nombre_usuario"]
        )
        
        return success_response(
            data=result,
            message="Producto restablecido exitosamente"
        )
        
    except HTTPException as e:
        return error_response(
            error="PRODUCT_RESTORE_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error restableciendo producto {product_id}: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.post("/{product_id}/upload-image", summary="Subir imagen de producto")
async def upload_product_image(
    product_id: int,
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    """
    Subir imagen para un producto
    
    - **product_id**: ID del producto
    - **file**: Archivo de imagen (JPG, PNG, máx 5MB)
    
    Requiere permisos de actualización de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "update")
        
        # Verificar que el producto existe
        await obtener_producto_por_id(product_id)
        
        # Validar archivo
        if not file.content_type.startswith('image/'):
            return error_response(
                error="INVALID_FILE_TYPE",
                message="Solo se permiten archivos de imagen",
                code=400
            )
        
        # Validar tamaño (5MB máximo)
        file_size = 0
        content = await file.read()
        file_size = len(content)
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return error_response(
                error="FILE_TOO_LARGE",
                message="El archivo es demasiado grande (máximo 5MB)",
                code=400
            )
        
        # Generar nombre único
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"product_{product_id}_{uuid.uuid4().hex}{file_extension}"
        
        # Crear directorio si no existe
        upload_dir = Path("static/images/products")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Guardar archivo
        file_path = upload_dir / unique_filename
        with open(file_path, "wb") as buffer:
            buffer.write(content)
        
        # Actualizar URL en producto

        url_foto = f"/static/images/products/{unique_filename}"
       
        from server.models.productos import ProductoUpdate
        update_data = ProductoUpdate(url_foto_producto=url_foto)
        
        await actualizar_producto(
            product_id,
            update_data,
            updated_by=current_user["user"]["id_usuario"],
            updated_by_name=current_user["user"]["nombre_usuario"]
        )
        
        return success_response(
            data={"url_foto": url_foto},
            message="Imagen subida exitosamente"
        )
       
    except HTTPException as e:
        return error_response(
            error="IMAGE_UPLOAD_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error subiendo imagen: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

@router.get("/validate/codigo/{codigo}", summary="Validar código único")
async def validate_codigo_producto(
   codigo: str,
   exclude_id: Optional[int] = Query(None, description="ID a excluir de la validación"),
   current_user = Depends(get_current_user)
):
   """
   Validar si un código de producto es único
   
   - **codigo**: Código a validar
   - **exclude_id**: ID de producto a excluir (para edición)
   
   Requiere permisos de lectura de productos
   """
   try:
       # Verificar permisos
       user_type = current_user["user"]["tipo_usuario"]
       check_product_permission(user_type, "read")
       
       is_unique = await verificar_codigo_producto_unico(codigo, exclude_id)
       
       return success_response(
           data={
               "codigo": codigo,
               "is_unique": is_unique,
               "available": is_unique
           },
           message="Validación completada"
       )
       
   except Exception as e:
       logger.error(f"Error validando código: {e}")
       return error_response(
           error="VALIDATION_ERROR",
           message="Error validando código",
           code=500
       )

@router.get("/export", summary="Exportar productos a Excel")
async def exportar_productos(
    current_user=Depends(get_current_user),
    estado: Optional[int] = Query(None, description="Filtro por estado (1=activo, 0=inactivo)"),
    tipo: Optional[str] = Query(None, description="Filtro por tipo de producto"),
):
    """
    Exportar productos a archivo Excel con filtros aplicados
    """
    try:
        # Debug: ver qué parámetros llegan
        logger.info(f"DEBUG PRODUCTOS EXPORT - estado: {estado}, tipo: {tipo}")
        logger.info(f"DEBUG PRODUCTOS EXPORT - user_type: {current_user['user']['tipo_usuario']}")
        
        # Verificar permisos
        check_product_permission(current_user["user"]["tipo_usuario"], "read")
        
        # Obtener productos con filtros
        productos_data = await obtener_productos(
            page=1,
            limit=10000,  # Obtener todos los productos
            estado=estado,
            tipo=tipo
        )
        
        productos = productos_data["data"]["productos"]
        
        # Crear workbook y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Definir headers
        headers = [
            "Código", "Nombre", "Descripción", "Tipo", "Categoría",
            "Proveedor", "Magnitud", "Stock Mínimo", "Stock Máximo", "Stock Crítico",
            "Costo Unitario", "Precio Referencial", "Ubicación Física", "Estado", "Fecha Creación"
        ]
        
        # Aplicar estilo a headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        for row, producto in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=producto.get("codigo_producto", ""))
            ws.cell(row=row, column=2, value=producto.get("nombre_producto", ""))
            ws.cell(row=row, column=3, value=producto.get("descripcion_producto", ""))
            ws.cell(row=row, column=4, value=producto.get("tipo_producto", ""))
            ws.cell(row=row, column=5, value=producto.get("categoria_producto", ""))
            ws.cell(row=row, column=6, value=producto.get("magnitud_producto", ""))
            ws.cell(row=row, column=7, value=producto.get("stock_minimo", 0))
            ws.cell(row=row, column=8, value=producto.get("stock_maximo", 0))
            ws.cell(row=row, column=9, value=producto.get("stock_critico", 0))
            ws.cell(row=row, column=10, value=producto.get("precio_unitario", 0))
            ws.cell(row=row, column=11, value=producto.get("ubicacion_fisica", ""))
            ws.cell(row=row, column=12, value="Activo" if producto.get("estado_producto", 1) == 1 else "Inactivo")
            ws.cell(row=row, column=13, value=str(producto.get("created_at", "")))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"productos_export_{timestamp}.xlsx"
        
        logger.info(f"Exportación de productos completada por usuario {current_user['user']['codigo_usuario']}")
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exportando productos: {e}")
        return error_response(
            error="EXPORT_ERROR",
            message="Error al exportar productos",
            code=500
        )

@router.post("/import", summary="Importar productos desde Excel")
async def importar_productos(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    """
    Importar productos masivamente desde archivo Excel
    
    - **file**: Archivo Excel (.xlsx, .xls) con los datos de productos
    
    El archivo debe contener las columnas:
    - codigo_producto (requerido)
    - nombre_producto (requerido) 
    - tipo_producto (requerido): insumo, repuesto, herramienta, otro
    - categoria_producto (opcional)
    - proveedor_producto (opcional)
    - costo_unitario (opcional)
    - precio_referencial (opcional)
    - ubicacion_fisica (opcional)
    - stock_minimo (opcional, default: 1)
    - stock_maximo (opcional, default: 100)
    - stock_critico (opcional, default: 0)
    - descripcion_producto (opcional)
    - magnitud_producto (opcional, default: UND)
    - requiere_lote (opcional, default: false)
    - dias_vida_util (opcional)
    
    Requiere permisos de creación de productos
    """
    try:
        # Verificar permisos
        user_type = current_user["user"]["tipo_usuario"]
        check_product_permission(user_type, "create")
        
        # Validar archivo
        if not file.content_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ]:
            return error_response(
                error="INVALID_FILE_TYPE",
                message="Solo se permiten archivos Excel (.xlsx, .xls)",
                code=400
            )
        
        # Validar tamaño (10MB máximo)
        file_content = await file.read()
        if len(file_content) > 10 * 1024 * 1024:  # 10MB
            return error_response(
                error="FILE_TOO_LARGE",
                message="El archivo es demasiado grande (máximo 10MB)",
                code=400
            )
        
        # Procesar importación
        resultado = await importar_productos_masivo(
            file_content,
            created_by=current_user["user"]["id_usuario"],
            created_by_name=current_user["user"]["nombre_usuario"]
        )
        
        # Preparar mensaje de respuesta
        if resultado.errores == 0:
            message = f"Importación exitosa: {resultado.exitosos} productos creados"
        elif resultado.exitosos == 0:
            message = f"Importación fallida: {resultado.errores} errores encontrados"
        else:
            message = f"Importación parcial: {resultado.exitosos} exitosos, {resultado.errores} errores"
        
        logger.info(f"Importación completada por usuario {current_user['user']['codigo_usuario']}: {message}")
        
        return success_response(
            data=resultado.dict(),
            message=message
        )
        
    except HTTPException as e:
        return error_response(
            error="IMPORT_FAILED",
            message=e.detail,
            code=e.status_code
        )
    except Exception as e:
        logger.error(f"Error importando productos: {e}")
        return error_response(
            error="INTERNAL_ERROR",
            message="Error interno del servidor",
            code=500
        )

        