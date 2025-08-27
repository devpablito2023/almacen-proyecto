# Script para crear archivo Excel de prueba
import openpyxl
import os

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

# Headers
headers = [
    "codigo_producto", "nombre_producto", "tipo_producto", "categoria_producto",
    "proveedor_producto", "costo_unitario", "precio_referencial", "ubicacion_fisica",
    "stock_minimo", "stock_maximo", "stock_critico", "descripcion_producto",
    "magnitud_producto", "requiere_lote", "dias_vida_util"
]

# Agregar headers
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Datos de ejemplo
data = [
    ["PROD001", "Aceite Hidráulico 20W", "insumo", "Lubricantes", "Proveedor A", 25.50, 30.00, "A1-001", 5, 50, 2, "Aceite hidráulico de alta calidad", "Litros", "NO", 365],
    ["PROD002", "Filtro de Aire", "repuesto", "Filtros", "Proveedor B", 15.00, 18.00, "B2-003", 3, 30, 1, "Filtro de aire para motores", "UND", "SI", 180],
    ["PROD003", "Tornillo M8x20", "herramienta", "Tornillería", "Proveedor C", 0.50, 0.75, "C3-010", 100, 1000, 50, "Tornillo métrico M8x20", "UND", "NO", None]
]

# Agregar datos
for row, product in enumerate(data, 2):
    for col, value in enumerate(product, 1):
        ws.cell(row=row, column=col, value=value)

# Crear directorio si no existe
excel_path = r"C:\Users\Ztrack2\Desktop\almacen-proyecto\test_import.xlsx"
print(f"Guardando archivo en: {excel_path}")

# Guardar archivo
wb.save(excel_path)
print("Archivo test_import.xlsx creado exitosamente")
